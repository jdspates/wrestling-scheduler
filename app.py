# app.py – Wrestling Scheduler – drag rows + rest gap warnings + scratches + manual matches 
import streamlit as st
import pandas as pd
import io
import random
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
import json
import os
import copy
from streamlit_js_eval import streamlit_js_eval
from datetime import datetime  # NEW: for autosave timestamp

from streamlit_sortables import sort_items  # drag-and-drop component

# ---------- Safe PatternFill import ----------
try:
    from openpyxl.styles import PatternFill
    _EXCEL_AVAILABLE = True
except Exception:
    _EXCEL_AVAILABLE = False

# ----------------------------------------------------------------------
# CONFIG & COLOR MAP
# ----------------------------------------------------------------------
CONFIG_FILE = "config.json"
AUTOSAVE_FILE = "autosave_meet.json"  # server-side autosave file

# 9-color palette that matches circle emojis
COLOR_MAP = {
    "red": "#FF0000",
    "orange": "#FF7F00",
    "yellow": "#FFD700",
    "green": "#008000",
    "blue": "#0000FF",
    "purple": "#800080",
    "brown": "#8B4513",
    "black": "#000000",
    "white": "#FFFFFF",
}

# Circle emojis – one per color, no duplicates
COLOR_ICON = {
    "red": "🔴",
    "orange": "🟠",
    "yellow": "🟡",
    "green": "🟢",
    "blue": "🔵",
    "purple": "🟣",
    "brown": "🟤",
    "black": "⚫",
    "white": "⚪",
}

DEFAULT_CONFIG = {
    "MIN_MATCHES": 2,
    "MAX_MATCHES": 4,
    "NUM_MATS": 4,
    "MAX_LEVEL_DIFF": 1,
    "WEIGHT_DIFF_FACTOR": 0.10,
    "MIN_WEIGHT_DIFF": 5.0,
    "REST_GAP": 4,  # minimum matches between bouts for same wrestler
    # TEAMS will be rebuilt from roster CSV once uploaded
    "TEAMS": []
}

# ----------------------------------------------------------------------
# ROSTER TEMPLATE (for new coaches)
# ----------------------------------------------------------------------
# Required columns the app expects:
# ["name", "team", "grade", "level", "weight", "early_matches", "scratch"]
# Optional columns:
# ["gender", "cross_gender_ok"]
#
# - gender: M/F (or variants like Male/Female/Boy/Girl – normalized in code)
# - cross_gender_ok: Y/N (or True/False-ish) – whether this wrestler allows cross-gender matches.
TEMPLATE_CSV = """name,team,grade,level,weight,early_matches,scratch,gender,cross_gender_ok
John Doe,Stillwater,7,1.0,70,N,N,M,Y
Jane Smith,Hastings,8,1.5,75,N,N,F,N
Ava Johnson,Woodbury,7,1.0,68,Y,N,F,Y
Mike Brown,Forest Lake,6,1.0,72,N,N,M,N
"""

# Load base config once (read-only default, e.g. from repo)
if os.path.exists(CONFIG_FILE):
    try:
        with open(CONFIG_FILE, "r") as f:
            loaded = json.load(f)
        if isinstance(loaded, dict):
            BASE_CONFIG = loaded
        else:
            BASE_CONFIG = DEFAULT_CONFIG
    except Exception:
        BASE_CONFIG = DEFAULT_CONFIG
else:
    BASE_CONFIG = DEFAULT_CONFIG

# ----------------------------------------------------------------------
# STYLES
# ----------------------------------------------------------------------
SORTABLE_STYLE = """
.sortable-component {
    background-color: transparent;
    border: none;
    padding: 0;
}
.sortable-container {
    background-color: transparent;
    border: none;
    box-shadow: none;
}
.sortable-container-header {
    display: none;
}
.sortable-container-body {
    background-color: transparent;
    padding: 0;
}
.sortable-item {
    background-color: #ffffff;
    color: #222 !important;
    border-radius: 4px;
    border: 1px solid #ddd;
    padding: 0 8px;
    margin-bottom: 3px;
    font-size: 0.82rem;
    font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    cursor: grab;

    height: 36px;
    display: flex;
    align-items: center;
}
.sortable-item:hover {
    background-color: #f7f7f7;
    color: #222 !important;
}
"""

# ----------------------------------------------------------------------
# SESSION STATE
# ----------------------------------------------------------------------
# Per-session CONFIG, cloned from BASE_CONFIG once
if "CONFIG" not in st.session_state:
    st.session_state.CONFIG = copy.deepcopy(BASE_CONFIG)

CONFIG = st.session_state.CONFIG  # convenience reference

for key in [
    "initialized", "bout_list", "mat_schedules", "suggestions",
    "active", "mat_order", "excel_bytes", "pdf_bytes", "coach_pdf_bytes",
    "roster", "manual_match_warning", "action_history"
]:
    if key not in st.session_state:
        if key in ["bout_list", "mat_schedules", "suggestions", "active", "action_history"]:
            st.session_state[key] = []
        elif key == "mat_order":
            st.session_state[key] = {}
        elif key in ["roster"]:
            st.session_state[key] = []
        elif key == "manual_match_warning":
            st.session_state[key] = ""
        else:
            st.session_state[key] = None

# version bump for sortable widgets so they refresh on add/remove/undo/scratches/color changes
if "sortable_version" not in st.session_state:
    st.session_state.sortable_version = 0

# versioned keys for file_uploaders so we can reset them cleanly
if "roster_uploader_version" not in st.session_state:
    st.session_state.roster_uploader_version = 0

# NEW: versioned key for JSON meet uploader (so Start Over can clear it)
if "state_json_uploader_version" not in st.session_state:
    st.session_state.state_json_uploader_version = 0

# NEW: confirmation flag for Start Over
if "reset_confirm" not in st.session_state:
    st.session_state.reset_confirm = False

# NEW: store last autosave time (for UI caption)
if "last_autosave_time" not in st.session_state:
    st.session_state.last_autosave_time = None

# NEW: map of bout_num -> overridden mat (for manual mat moves)
if "mat_overrides" not in st.session_state:
    st.session_state.mat_overrides = {}

# ----------------------------------------------------------------------
# GENDER HELPERS (NEW)
# ----------------------------------------------------------------------
def _parse_gender(val):
    """Normalize gender value to 'M', 'F', or None."""
    if pd.isna(val):
        return None
    s = str(val).strip().upper()
    if s in ["M", "MALE", "B", "BOY"]:
        return "M"
    if s in ["F", "FEMALE", "G", "GIRL"]:
        return "F"
    return None  # unknown/unset


def _parse_cross_gender_ok(val):
    """
    Normalize cross-gender flag to bool.
    Default True for backwards compatibility (no column = no restriction).
    """
    if pd.isna(val):
        return True
    s = str(val).strip().upper()
    return s in ["Y", "YES", "TRUE", "T", "1"]


def genders_compatible(w1, w2):
    """
    Gender matching rule:

    - If either wrestler has no gender recorded, allow (no gender constraint).
    - If both same gender, always allow.
    - If genders differ, allow only if BOTH have cross_gender_ok = True.
    """
    g1 = w1.get("gender")
    g2 = w2.get("gender")

    # If either missing/unknown, don't enforce gender constraint
    if not g1 or not g2:
        return True

    # Same gender always OK
    if g1 == g2:
        return True

    # Cross-gender only if BOTH have cross_gender_ok = True
    c1 = w1.get("cross_gender_ok", True)
    c2 = w2.get("cross_gender_ok", True)
    return bool(c1 and c2)

# ----------------------------------------------------------------------
# CORE LOGIC
# ----------------------------------------------------------------------
def is_compatible(w1, w2):
    """
    Base compatibility check:
      - different teams
      - avoid 5th vs 7th/8th graders
      - respect gender preferences (NEW)
    """
    return (
        w1["team"] != w2["team"]
        and not (
            (w1["grade"] == 5 and w2["grade"] in [7, 8]) or
            (w2["grade"] == 5 and w1["grade"] in [7, 8])
        )
        and genders_compatible(w1, w2)
    )


def max_weight_diff(w):
    return max(CONFIG["MIN_WEIGHT_DIFF"], w * CONFIG["WEIGHT_DIFF_FACTOR"])


def matchup_score(w1, w2):
    return round(
        abs(w1["weight"] - w2["weight"]) +
        abs(w1["level"] - w2["level"]) * 10, 1
    )


def generate_initial_matchups(active):
    bouts = set()
    for level in sorted({w["level"] for w in active}, reverse=True):
        group = [w for w in active if w["level"] == level]
        while True:
            added = False
            random.shuffle(group)
            for w in group:
                if len(w["match_ids"]) >= CONFIG["MAX_MATCHES"]:
                    continue
                opps = [
                    o for o in active
                    if o["id"] not in w["match_ids"]
                    and o["id"] != w["id"]
                    and len(o["match_ids"]) < CONFIG["MAX_MATCHES"]
                    and is_compatible(w, o)
                    and abs(w["weight"] - o["weight"]) <= \
                        min(max_weight_diff(w["weight"]), max_weight_diff(o["weight"]))
                    and abs(w["level"] - o["level"]) <= CONFIG["MAX_LEVEL_DIFF"]
                ]
                if not opps:
                    continue
                best = min(opps, key=lambda o: matchup_score(w, o))
                w["match_ids"].append(best["id"])
                best["match_ids"].append(w["id"])
                bouts.add(frozenset({w["id"], best["id"]}))
                added = True
                break
            if not added:
                break

    bout_list = []
    for idx, b in enumerate(bouts, 1):
        w1 = next(w for w in active if w["id"] == list(b)[0])
        w2 = next(w for w in active if w["id"] == list(b)[1])
        bout_list.append({
            "bout_num": idx,
            "w1_id": w1["id"], "w1_name": w1["name"], "w1_team": w1["team"],
            "w1_level": w1["level"], "w1_weight": w1["weight"],
            "w1_grade": w1["grade"], "w1_early": w1["early"],
            "w2_id": w2["id"], "w2_name": w2["name"], "w2_team": w2["team"],
            "w2_level": w2["level"], "w2_weight": w2["weight"],
            "w2_grade": w2["grade"], "w2_early": w2["early"],
            "score": matchup_score(w1, w2),
            "avg_weight": (w1["weight"] + w2["weight"]) / 2,
            "is_early": w1["early"] or w2["early"],
            "manual": ""
        })
    return bout_list


def build_suggestions(active, bout_list):
    """
    Suggest additional matches for wrestlers under MIN_MATCHES.
    Now respects gender preferences via genders_compatible().
    """
    under = [w for w in active if len(w["match_ids"]) < CONFIG["MIN_MATCHES"]]
    sugg = []
    for w in under:
        opps = [o for o in active if o["id"] not in w["match_ids"] and o["id"] != w["id"]]
        opps = [
            o for o in opps
            if genders_compatible(w, o)  # NEW gender filter
            and abs(w["weight"] - o["weight"]) <= \
                min(max_weight_diff(w["weight"]), max_weight_diff(o["weight"]))
            and abs(w["level"] - o["level"]) <= CONFIG["MAX_LEVEL_DIFF"]
        ]
        if not opps:
            # Fallback – any opponent not yet matched, but still gender-compatible if possible
            opps = [
                o for o in active
                if o["id"] not in w["match_ids"]
                and o["id"] != w["id"]
                and genders_compatible(w, o)
            ]
        for o in sorted(opps, key=lambda o: matchup_score(w, o))[:3]:
            sugg.append({
                "wrestler": w["name"], "team": w["team"],
                "level": w["level"], "weight": w["weight"],
                "current": len(w["match_ids"]),
                "vs": o["name"], "vs_team": o["team"],
                "vs_current": len(o["match_ids"]),
                "vs_level": o["level"], "vs_weight": o["weight"],
                "score": matchup_score(w, o),
                "_w_id": w["id"], "_o_id": o["id"]
            })
    return sugg


def generate_mat_schedule(bout_list, gap=4):
    """Base scheduling algorithm (ignores manual ordering, respects manual removals)."""
    valid = [b for b in bout_list if b["manual"] != "Manually Removed"]
    valid = sorted(valid, key=lambda x: x["avg_weight"])  # Light to heavy

    per_mat = len(valid) // CONFIG["NUM_MATS"]
    extra = len(valid) % CONFIG["NUM_MATS"]
    mats = []
    start = 0
    for i in range(CONFIG["NUM_MATS"]):
        end = start + per_mat + (1 if i < extra else 0)
        mats.append(valid[start:end])
        start = end

    schedules = []
    last_slot = {}

    for mat_num, mat_bouts in enumerate(mats, 1):
        early_bouts = [b for b in mat_bouts if b["is_early"]]
        non_early_bouts = [b for b in mat_bouts if not b["is_early"]]
        total_slots = len(mat_bouts)
        first_half_end = (total_slots + 1) // 2
        slot = 1
        scheduled = []
        first_half_wrestlers = set()

        # First early match if possible
        first_early = None
        for b in early_bouts:
            l1 = last_slot.get(b["w1_id"], -100)
            l2 = last_slot.get(b["w2_id"], -100)
            if l1 < 0 and l2 < 0:
                first_early = b
                break
        if first_early:
            early_bouts.remove(first_early)
            scheduled.append((1, first_early))
            last_slot[first_early["w1_id"]] = 1
            last_slot[first_early["w2_id"]] = 1
            first_half_wrestlers.update([first_early["w1_id"], first_early["w2_id"]])
            slot = 2

        # More early matches in first half
        while early_bouts and len(scheduled) < first_half_end:
            best = None
            best_score = -float("inf")
            for b in early_bouts:
                if b["w1_id"] in first_half_wrestlers or b["w2_id"] in first_half_wrestlers:
                    continue
                l1 = last_slot.get(b["w1_id"], -100)
                l2 = last_slot.get(b["w2_id"], -100)
                if l1 >= slot - 1 or l2 >= slot - 1:
                    continue
                score = min(slot - l1 - 1, slot - l2 - 1)
                if score > best_score:
                    best_score = score
                    best = b
            if best is None:
                break
            early_bouts.remove(best)
            scheduled.append((slot, best))
            last_slot[best["w1_id"]] = slot
            last_slot[best["w2_id"]] = slot
            first_half_wrestlers.update([best["w1_id"], best["w2_id"]])
            slot += 1

        # Remaining matches
        remaining = non_early_bouts + early_bouts
        while remaining:
            best = None
            best_gap = -1
            for b in remaining:
                l1 = last_slot.get(b["w1_id"], -100)
                l2 = last_slot.get(b["w2_id"], -100)
                if l1 >= slot - gap or l2 >= slot - gap:
                    continue
                gap_val = min(slot - l1 - 1, slot - l2 - 1)
                if gap_val > best_gap:
                    best_gap = gap_val
                    best = b
            if best is None and remaining:
                best = remaining[0]
            remaining.remove(best)
            scheduled.append((slot, best))
            last_slot[best["w1_id"]] = slot
            last_slot[best["w2_id"]] = slot
            slot += 1

        for s, b in scheduled:
            schedules.append({
                "mat": mat_num,
                "slot": s,
                "bout_num": b["bout_num"],
                "w1": f"{b['w1_name']} ({b['w1_team']})",
                "w2": f"{b['w2_name']} ({b['w2_team']})",
                "w1_team": b["w1_team"],
                "w2_team": b["w2_team"],
                "is_early": b["is_early"]
            })

    for mat_num in range(1, CONFIG["NUM_MATS"] + 1):
        mat_entries = [m for m in schedules if m["mat"] == mat_num]
        mat_entries.sort(key=lambda x: x["slot"])
        for idx, entry in enumerate(mat_entries, 1):
            entry["mat_bout_num"] = idx

    return schedules


def apply_mat_order_to_global_schedule():
    """
    Take the base schedule, then:
      - apply any mat overrides (bout -> mat),
      - reorder each mat according to st.session_state.mat_order,
      - recompute slot + mat_bout_num so exports and previews match the dragged order.
    """
    rest_gap = CONFIG.get("REST_GAP", 4)
    base = generate_mat_schedule(st.session_state.bout_list, gap=rest_gap)

    # NEW: apply mat overrides (manual moves)
    overrides = st.session_state.get("mat_overrides", {})
    if overrides:
        for e in base:
            override_mat = overrides.get(e["bout_num"])
            if override_mat:
                e["mat"] = override_mat

    schedules = []

    for mat in range(1, CONFIG["NUM_MATS"] + 1):
        entries = [e for e in base if e["mat"] == mat]
        order = st.session_state.mat_order.get(mat)

        if order:
            entries_sorted = sorted(
                entries,
                key=lambda e: (
                    order.index(e["bout_num"])
                    if e["bout_num"] in order
                    else len(order) + e["slot"]
                )
            )
        else:
            entries_sorted = sorted(entries, key=lambda e: e["slot"])

        for idx, e in enumerate(entries_sorted, start=1):
            e["slot"] = idx
            e["mat_bout_num"] = idx
            schedules.append(e)

    return schedules


def compute_rest_conflicts(schedule, min_gap):
    """
    Given a flat schedule (list of entries with mat, slot, bout_num),
    find wrestlers who have matches too close together (slot difference < min_gap).
    Returns a list of dicts with details for display.
    """
    appearances = {}

    for e in schedule:
        b = next(x for x in st.session_state.bout_list if x["bout_num"] == e["bout_num"])

        for w_id, name, team in [
            (b["w1_id"], b["w1_name"], b["w1_team"]),
            (b["w2_id"], b["w2_name"], b["w2_team"]),
        ]:
            if w_id not in appearances:
                appearances[w_id] = {
                    "name": name,
                    "team": team,
                    "matches": []
                }
            appearances[w_id]["matches"].append((e["mat"], e["slot"], e["bout_num"]))

    conflicts = []

    for w_id, info in appearances.items():
        by_mat = {}
        for mat, slot, bout_num in info["matches"]:
            by_mat.setdefault(mat, []).append((slot, bout_num))

        for mat, matches in by_mat.items():
            matches.sort(key=lambda x: x[0])  # sort by slot
            for (slot1, bout1), (slot2, bout2) in zip(matches, matches[1:]):
                gap = slot2 - slot1
                if gap < min_gap:
                    conflicts.append({
                        "wrestler_id": w_id,
                        "wrestler": info["name"],
                        "team": info["team"],
                        "mat": mat,
                        "slot1": slot1,
                        "slot2": slot2,
                        "bout1": bout1,
                        "bout2": bout2,
                        "gap": gap,
                    })

    return conflicts


def compute_multi_mat_assignments(schedule):
    """
    Find wrestlers who are scheduled on more than one mat.
    Returns a list of dicts:
      {
        wrestler_id,
        name,
        team,
        mats: sorted list of mats,
        matches: list of {mat, slot, bout_num}
      }
    """
    appearances = {}

    for e in schedule:
        b = next(x for x in st.session_state.bout_list if x["bout_num"] == e["bout_num"])

        for w_id, name, team in [
            (b["w1_id"], b["w1_name"], b["w1_team"]),
            (b["w2_id"], b["w2_name"], b["w2_team"]),
        ]:
            if w_id not in appearances:
                appearances[w_id] = {
                    "name": name,
                    "team": team,
                    "matches": [],  # list of {mat, slot, bout_num}
                }
            appearances[w_id]["matches"].append({
                "mat": e["mat"],
                "slot": e["slot"],          # this is the visible slot on that mat
                "bout_num": e["bout_num"],
            })

    multi = []
    for w_id, info in appearances.items():
        mats = sorted({m["mat"] for m in info["matches"]})
        if len(mats) > 1:
            multi.append({
                "wrestler_id": w_id,
                "name": info["name"],
                "team": info["team"],
                "mats": mats,
                # sort matches nicely by mat, then slot
                "matches": sorted(info["matches"], key=lambda x: (x["mat"], x["slot"])),
            })

    return multi

def _short_name(full_name: str) -> str:
    """Turn 'Brady Stebbins' into 'B. Stebbins'."""
    if not full_name:
        return ""
    parts = str(full_name).split()
    if len(parts) == 1:
        return parts[0]
    first_initial = parts[0][0].upper() + "."
    last = parts[-1].capitalize()
    return f"{first_initial} {last}"


def _team_abbrev(team_name: str) -> str:
    """Turn 'Forest Lake' into 'FL', 'Stillwater' into 'STI', etc."""
    if not team_name:
        return ""
    parts = [p for p in str(team_name).split() if p]
    if len(parts) == 1:
        return parts[0][:3].upper()
    # First two initials, e.g. Forest Lake -> FL, East Ridge -> ER
    return "".join(p[0].upper() for p in parts[:2])

def generate_coach_packets_pdf(full_schedule):
    """
    Build a PDF with one page per team (landscape).
    Each page lists all active wrestlers on that team and ALL of their matches
    (across mats), with dynamic Match 1 / Match 2 / ... columns.

    Columns: Wrestler | Wt | Match 1 | Match 2 | ...
    Match cell example: "M1 S18: B. Stebbins (FL)"
    """
    from reportlab.lib.pagesizes import landscape, letter  # in case not imported at top

    buf = io.BytesIO()
    # LANDSCAPE page setup
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
    page_width, page_height = landscape(letter)
    avail_width = page_width - doc.leftMargin - doc.rightMargin

    elements = []
    styles = getSampleStyleSheet()

    # --- Small helpers for abbreviations ---------------------------------
    def abbreviate_team(team: str) -> str:
        """
        Make a short team code like:
        "Forest Lake" -> "FL"
        "Stillwater"  -> "S"
        "New Prague"  -> "NP"
        """
        if not team:
            return ""
        parts = [p for p in str(team).split() if p.strip()]
        initials = "".join(p[0].upper() for p in parts)
        return initials[:3]

    def abbreviate_name(full_name: str) -> str:
        """
        "Brandon Stebbins" -> "B. Stebbins"
        "Aiden" -> "Aiden"
        "Mary Ann Smith" -> "M. Smith"
        """
        if not full_name:
            return ""
        parts = [p for p in str(full_name).split() if p.strip()]
        if len(parts) == 1:
            return parts[0]
        first = parts[0]
        last = parts[-1]
        return f"{first[0].upper()}. {last}"

    # Map wrestler_id -> wrestler record (only active wrestlers)
    active = st.session_state.get("active", [])
    wrestler_by_id = {w["id"]: w for w in active}

    # Build per-wrestler match info
    # key: (team, wrestler_id) -> {
    #   "name", "team", "weight", "matches": [ {mat, slot, opp_name, opp_team} ]
    # }
    packets = {}

    for e in full_schedule:
        # Look up the bout behind this schedule entry
        try:
            b = next(x for x in st.session_state.bout_list if x["bout_num"] == e["bout_num"])
        except StopIteration:
            continue

        # Add both sides (w1 and w2) to their own team's packet
        for side, opp_side in (("w1", "w2"), ("w2", "w1")):
            wid = b.get(f"{side}_id")
            w = wrestler_by_id.get(wid)
            if not w:
                continue

            team = w["team"]
            key = (team, wid)
            if key not in packets:
                packets[key] = {
                    "name": w["name"],
                    "team": team,
                    "weight": w["weight"],
                    "matches": []
                }

            packets[key]["matches"].append({
                "mat": e["mat"],
                "slot": e["slot"],
                "opp_name": b.get(f"{opp_side}_name"),
                "opp_team": b.get(f"{opp_side}_team"),
            })

    # Group wrestlers by team
    team_to_wrestlers = {}
    for (team, wid), rec in packets.items():
        team_to_wrestlers.setdefault(team, []).append(rec)

    if not team_to_wrestlers:
        # No matches / no active wrestlers
        elements.append(Paragraph("No coach packets to generate (no matches found).", styles["Normal"]))
        doc.build(elements)
        return buf.getvalue()

    first_team = True

    for team, wrestlers in sorted(team_to_wrestlers.items()):
        # Sort wrestlers (light → heavy, then name)
        wrestlers.sort(key=lambda r: (r["weight"], r["name"]))

        # How many match columns do we need?
        max_matches = max((len(r["matches"]) for r in wrestlers), default=0)

        # Headers: Wrestler / Wt / Match 1 / Match 2 / ... / Match N
        headers = ["Wrestler", "Wt"] + [
            f"Match {i + 1}" for i in range(max_matches)
        ]

        table_data = [headers]

        # Build each row with exactly max_matches match cells
        for r in wrestlers:
            row = [
                r["name"],
                f"{r['weight']:.0f}",
            ]

            # Add each match cell with abbreviated text:
            # "M1 S18: B. Stebbins (FL)"
            for m in r["matches"]:
                opp_name = abbreviate_name(m["opp_name"])
                opp_team_short = abbreviate_team(m["opp_team"])
                cell = f"M{m['mat']} S{m['slot']}: {opp_name} ({opp_team_short})"
                row.append(cell)

            # Pad with empty strings so every row has the same number of columns
            while len(row) < 2 + max_matches:
                row.append("")

            table_data.append(row)

        # Column widths:
        # - First 2 columns fixed
        # - Remaining columns share whatever width is left on the page
        # You can tweak these to taste if things still feel tight.
        fixed_widths = [1.8 * inch, 0.45 * inch]  # Wrestler, Wt

        # Make sure we don't go negative even if margins change
        remaining_width = max(avail_width - sum(fixed_widths), 2.0 * inch)

        if max_matches > 0:
            match_width = remaining_width / max_matches
            col_widths = fixed_widths + [match_width] * max_matches
        else:
            col_widths = fixed_widths

        table = Table(table_data, colWidths=col_widths)

        style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.lightgrey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),  # smaller text so long match info fits
        ])
        table.setStyle(style)

        # Page break between teams
        if not first_team:
            elements.append(PageBreak())
        first_team = False

        elements.append(Paragraph(f"{team} – Coach Packet", styles["Title"]))
        elements.append(Spacer(1, 12))
        elements.append(table)

    doc.build(elements)
    return buf.getvalue()

# ----------------------------------------------------------------------
# HELPERS (undo + color dots)
# ----------------------------------------------------------------------
def color_dot_hex(hex_color: str) -> str:
    """Return a small HTML circle for the given hex color (for legends / HTML tables)."""
    if not hex_color:
        return ""
    return (
        "<span style='display:inline-block;width:12px;height:12px;"
        f"border-radius:50%;background:{hex_color};margin-right:6px;'></span>"
    )


def push_action(action: dict):
    """Record an action so it can be undone later."""
    if "action_history" not in st.session_state:
        st.session_state.action_history = []
    st.session_state.action_history.append(action)


def _undo_remove(bout_num: int):
    """Undo a previously removed bout."""
    try:
        b = next(
            x for x in st.session_state.bout_list
            if x["bout_num"] == bout_num and x.get("manual") == "Manually Removed"
        )
    except StopIteration:
        st.info("Removed bout not found; nothing to undo.")
        return

    b["manual"] = ""
    w1 = next(w for w in st.session_state.active if w["id"] == b["w1_id"])
    w2 = next(w for w in st.session_state.active if w["id"] == b["w2_id"])

    if b["w2_id"] not in w1["match_ids"]:
        w1["match_ids"].append(b["w2_id"])
    if b["w1_id"] not in w2["match_ids"]:
        w2["match_ids"].append(b["w1_id"])

    st.session_state.bout_list.sort(key=lambda x: x["avg_weight"])
    st.session_state.mat_order = {}   # keep behavior: layout recalculated
    st.session_state.suggestions = build_suggestions(st.session_state.active, st.session_state.bout_list)
    st.session_state.excel_bytes = None
    st.session_state.pdf_bytes = None
    st.session_state.sortable_version += 1
    st.success("Undo: restored last removed bout.")


def _undo_drag(previous_mat_order: dict):
    """Undo a drag/reorder by restoring previous mat_order snapshot."""
    st.session_state.mat_order = {
        m: order.copy() for m, order in previous_mat_order.items()
    }
    st.session_state.excel_bytes = None
    st.session_state.pdf_bytes = None
    st.session_state.sortable_version += 1
    st.success("Undo: last drag / reorder reverted.")


def _undo_manual_add(bout_num: int):
    """Undo a manually-added match."""
    try:
        b = next(x for x in st.session_state.bout_list if x["bout_num"] == bout_num)
    except StopIteration:
        st.info("Manual match already removed; nothing to undo.")
        return

    w1 = next(w for w in st.session_state.active if w["id"] == b["w1_id"])
    w2 = next(w for w in st.session_state.active if w["id"] == b["w2_id"])

    if b["w2_id"] in w1["match_ids"]:
        w1["match_ids"].remove(b["w2_id"])
    if b["w1_id"] in w2["match_ids"]:
        w2["match_ids"].remove(b["w1_id"])

    # Remove bout from bout_list
    st.session_state.bout_list = [
        x for x in st.session_state.bout_list if x["bout_num"] != bout_num
    ]

    # Strip from any mat_order lists
    for mat, order in st.session_state.mat_order.items():
        st.session_state.mat_order[mat] = [bn for bn in order if bn != bout_num]

    st.session_state.suggestions = build_suggestions(st.session_state.active, st.session_state.bout_list)
    st.session_state.excel_bytes = None
    st.session_state.pdf_bytes = None
    st.session_state.sortable_version += 1
    st.success("Undo: manual match removed.")


def _undo_suggest_add(bout_nums: list[int]):
    """Undo a batch of suggested matches that were added at once."""
    bout_nums_set = set(bout_nums)

    # Clean wrestler match_ids
    for b in list(st.session_state.bout_list):
        if b["bout_num"] in bout_nums_set:
            w1 = next(w for w in st.session_state.active if w["id"] == b["w1_id"])
            w2 = next(w for w in st.session_state.active if w["id"] == b["w2_id"])

            if b["w2_id"] in w1["match_ids"]:
                w1["match_ids"].remove(b["w2_id"])
            if b["w1_id"] in w2["match_ids"]:
                w2["match_ids"].remove(b["w1_id"])

    # Remove bouts
    st.session_state.bout_list = [
        x for x in st.session_state.bout_list if x["bout_num"] not in bout_nums_set
    ]

    # Strip from mat_order
    for mat, order in st.session_state.mat_order.items():
        st.session_state.mat_order[mat] = [
            bn for bn in order if bn not in bout_nums_set
        ]

    st.session_state.suggestions = build_suggestions(st.session_state.active, st.session_state.bout_list)
    st.session_state.excel_bytes = None
    st.session_state.pdf_bytes = None
    st.session_state.sortable_version += 1
    st.success("Undo: suggested matches removed.")


def _undo_scratch_update(snapshot: dict):
    """Undo a scratches update by restoring a saved snapshot."""
    # Restore from snapshot using deep copies so we don't share references
    st.session_state.roster = copy.deepcopy(snapshot["roster"])
    st.session_state.active = copy.deepcopy(snapshot["active"])
    st.session_state.bout_list = copy.deepcopy(snapshot["bout_list"])
    st.session_state.suggestions = copy.deepcopy(snapshot["suggestions"])
    st.session_state.mat_order = copy.deepcopy(snapshot["mat_order"])
    st.session_state.mat_overrides = copy.deepcopy(snapshot.get("mat_overrides", {}))

    # The Pre-Meet Scratches widget will rebuild its selection from
    # the restored roster (w['scratch']) on the next run.
    st.session_state.excel_bytes = None
    st.session_state.pdf_bytes = None
    st.session_state.sortable_version += 1
    st.success("Undo: scratches and schedule restored.")


def undo_last_action():
    """Pop the last action off the history and undo it."""
    history = st.session_state.get("action_history", [])
    if not history:
        st.info("No actions to undo yet.")
        return

    action = history.pop()
    t = action.get("type")

    if t == "remove":
        _undo_remove(action["bout_num"])
    elif t == "drag":
        _undo_drag(action["previous_mat_order"])
    elif t == "manual_add":
        _undo_manual_add(action["bout_num"])
    elif t == "suggest_add":
        _undo_suggest_add(action["bout_nums"])
    elif t == "scratch_update":
        _undo_scratch_update(action["snapshot"])
    else:
        st.info("Nothing to undo.")
        return

    st.rerun()


def remove_bout(bout_num: int):
    """Mark bout as manually removed, update wrestler match_ids, trim from mat_order."""
    try:
        b = next(x for x in st.session_state.bout_list if x["bout_num"] == bout_num)
    except StopIteration:
        return
    if b.get("manual") == "Manually Removed":
        return

    b["manual"] = "Manually Removed"
    w1 = next(w for w in st.session_state.active if w["id"] == b["w1_id"])
    w2 = next(w for w in st.session_state.active if w["id"] == b["w2_id"])
    if b["w2_id"] in w1["match_ids"]:
        w1["match_ids"].remove(b["w2_id"])
    if b["w1_id"] in w2["match_ids"]:
        w2["match_ids"].remove(b["w1_id"])

    # Push remove action to history
    push_action({"type": "remove", "bout_num": bout_num})

    # Remove bout from any mat orders
    for mat, order in st.session_state.mat_order.items():
        if bout_num in order:
            order.remove(bout_num)

    st.session_state.suggestions = build_suggestions(st.session_state.active, st.session_state.bout_list)
    st.session_state.excel_bytes = None
    st.session_state.pdf_bytes = None

    st.session_state.sortable_version += 1
    st.rerun()


def validate_roster_df(df: pd.DataFrame):
    """Return list of error messages if roster has issues; empty list if OK."""
    errors = []
    # NOTE: no 'id' column required now – IDs are generated internally
    # gender and cross_gender_ok are OPTIONAL
    required = ["name", "team", "grade", "level", "weight", "early_matches", "scratch"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        errors.append("Missing columns: " + ", ".join(missing))
        return errors

    if df.empty:
        errors.append("Roster file is empty (no wrestlers found).")

    # Basic numeric checks
    for col in ["grade", "level", "weight"]:
        bad = pd.to_numeric(df[col], errors="coerce").isna()
        if bad.any():
            bad_vals = df.loc[bad, col].astype(str).unique().tolist()
            errors.append(f"Column '{col}' has non-numeric values: {bad_vals}")

    return errors

# ----------------------------------------------------------------------
# WIDGET RESET HELPER (for restored meets)
# ----------------------------------------------------------------------
def reset_setting_widgets():
    """
    Clear sidebar / search / team color widget keys so that on the next run,
    the widgets use the restored CONFIG and team colors instead of stale values.
    """
    # Numeric / slider settings
    for key in [
        "min_matches",
        "max_matches",
        "num_mats",
        "max_level_diff",
        "min_weight_diff",
        "weight_factor",
        "rest_gap",
    ]:
        st.session_state.pop(key, None)

    # Wrestler search box
    st.session_state.pop("wrestler_search", None)

    # Team color selectboxes: keys look like "color_0", "color_1", ...
    color_keys = [k for k in list(st.session_state.keys()) if k.startswith("color_")]
    for k in color_keys:
        st.session_state.pop(k, None)

# ----------------------------------------------------------------------
# SNAPSHOT SAVE / LOAD HELPERS (JSON)
# ----------------------------------------------------------------------
def build_meet_snapshot():
    """Return a JSON-serializable snapshot of the current meet."""
    return {
        "CONFIG": st.session_state.CONFIG,
        "roster": st.session_state.get("roster", []),
        "active": st.session_state.get("active", []),
        "bout_list": st.session_state.get("bout_list", []),
        "suggestions": st.session_state.get("suggestions", []),
        "mat_order": st.session_state.get("mat_order", {}),
    }


def restore_meet_from_snapshot(data: dict):
    """Restore a meet snapshot into session_state."""
    # Load CONFIG from snapshot
    st.session_state.CONFIG = data.get("CONFIG", DEFAULT_CONFIG)
    cfg = st.session_state.CONFIG

    # Clear widget state so sidebar & colors pick up restored CONFIG/TEAMS
    reset_setting_widgets()

    # NEW: explicitly sync widget-backed keys to loaded CONFIG
    st.session_state["min_matches"] = cfg["MIN_MATCHES"]
    st.session_state["max_matches"] = cfg["MAX_MATCHES"]
    st.session_state["num_mats"] = cfg["NUM_MATS"]
    st.session_state["max_level_diff"] = cfg["MAX_LEVEL_DIFF"]
    st.session_state["min_weight_diff"] = cfg["MIN_WEIGHT_DIFF"]
    st.session_state["weight_factor"] = cfg["WEIGHT_DIFF_FACTOR"]
    st.session_state["rest_gap"] = cfg.get("REST_GAP", 4)

    st.session_state.roster = data.get("roster", [])
    st.session_state.active = data.get("active", [])
    st.session_state.bout_list = data.get("bout_list", [])
    st.session_state.suggestions = data.get("suggestions", [])
    st.session_state.mat_order = data.get("mat_order", {})
    st.session_state.excel_bytes = None
    st.session_state.pdf_bytes = None
    st.session_state.initialized = bool(st.session_state.roster)
    st.session_state.sortable_version += 1  # refresh drag widgets
    st.session_state.action_history = []  # clear undo history on restore


def autosave_meet():
    try:
        snapshot = build_meet_snapshot()
        with open(AUTOSAVE_FILE, "w", encoding="utf-8") as f:
            json.dump(snapshot, f)

        local_time = streamlit_js_eval(
            js_expressions="new Date().toLocaleTimeString([], {hour: 'numeric', minute: '2-digit'})",
            key="local_time_key"
        )

        if local_time:
            st.session_state["last_autosave_time"] = local_time

    except Exception:
        pass

# ----------------------------------------------------------------------
# STREAMLIT APP LAYOUT
# ----------------------------------------------------------------------
st.set_page_config(page_title="Wrestling Scheduler", layout="wide")

st.markdown(
    """
<style>
    div[data-testid="stExpander"] > div > div { padding:0 !important; margin:0 !important; }
    div[data-testid="stVerticalBlock"] > div { gap:0 !important; }
    .block-container { padding:2rem 1rem !important; max-width:1200px !important; margin:0 auto !important; }
    .main .block-container { padding-left:2rem !important; padding-right:2rem !important; }
    h1 { margin-top:0 !important; }
    .stSidebar .stButton > button {
        padding: 0.5rem 1rem !important;
        height: auto !important;
        min-width: auto !important;
    }
    .stTextInput > div > div > input { border-radius: 6px !important; }
    .stTextInput > div > div > button {
        background: transparent !important;
        border: none !important;
        color: #888 !important;
    }
</style>
""",
    unsafe_allow_html=True,
)
st.markdown(f"<style>{SORTABLE_STYLE}</style>", unsafe_allow_html=True)

st.title("Wrestling Meet Scheduler")
st.caption("Upload roster → Generate → Edit → Download. **No data stored.**")

# ---- STEP 1: DOWNLOAD ROSTER TEMPLATE ----
st.markdown("### Step 1 – Download roster template (CSV)")
st.markdown(
    "Download the example file, add your wrestlers, save it as a `.csv`, "
    "then upload it in Step 2 below."
)

st.download_button(
    label="⬇️ Download roster template CSV",
    data=TEMPLATE_CSV.encode("utf-8"),
    file_name="roster_template.csv",
    mime="text/csv",
    use_container_width=False,
)

st.markdown("---")

# ---- STEP 2: UPLOAD ROSTER ----
st.markdown("### Step 2 – Upload your completed `roster.csv`")

uploaded = st.file_uploader(
    "Upload your roster.csv file",
    type="csv",
    key=f"roster_csv_uploader_v{st.session_state.roster_uploader_version}",
)

# Process upload once per meet
if uploaded and not st.session_state.initialized:
    try:
        df = pd.read_csv(uploaded)

        # Validate first
        validation_errors = validate_roster_df(df)
        if validation_errors:
            for msg in validation_errors:
                st.error(msg)
            st.stop()

        wrestlers = df.to_dict("records")

        # Generate internal integer IDs
        for idx, w in enumerate(wrestlers, start=1):
            w["id"] = idx
            w["grade"] = int(w["grade"])
            w["level"] = float(w["level"])
            w["weight"] = float(w["weight"])
            w["early"] = (
                str(w["early_matches"]).strip().upper() == "Y"
                or w["early_matches"] in [1, True]
            )
            w["scratch"] = (
                str(w["scratch"]).strip().upper() == "Y"
                or w["scratch"] in [1, True]
            )

            # --- NEW: gender + cross_gender_ok (optional columns) ---
            w["gender"] = _parse_gender(w.get("gender", None))
            w["cross_gender_ok"] = _parse_cross_gender_ok(w.get("cross_gender_ok", None))

            w["match_ids"] = []

        st.session_state.roster = wrestlers
        st.session_state.active = [w for w in wrestlers if not w["scratch"]]
        st.session_state.bout_list = generate_initial_matchups(st.session_state.active)
        st.session_state.suggestions = build_suggestions(st.session_state.active, st.session_state.bout_list)
        st.session_state.initialized = True
        st.session_state.action_history = []

        st.success(
            f"Roster loaded ({len(wrestlers)} wrestlers, "
            f"{len({w['team'] for w in wrestlers})} teams) and matchups generated!"
        )
    except Exception as e:
        st.error(f"Error loading roster: {e}")

# ----------------------------------------------------------------------
# ADVANCED OPTIONS – START OVER + SAVE / LOAD MEET + MERGE ROSTERS
# ----------------------------------------------------------------------
with st.expander("Advanced options (Start Over, Save / Load meet / Merge CSV Roster Files)", expanded=False):
    st.caption(
        "Optional tools for resetting this meet, saving/loading meet files, "
        "or merging multiple team roster CSVs into one file. "
        "Most coaches won't need these every time."
    )

    # ----- Start Over / Load New Roster -----
    if st.session_state.get("initialized") and st.session_state.get("roster"):

        st.markdown("##### Start Over / Load New Roster")

        # Show either Start Over button OR confirmation UI, never both
        if not st.session_state.get("reset_confirm", False):
            # Primary Start Over button – toggles confirmation mode
            if st.button(
                "🔄 Start Over / Load New Roster",
                help="Clear current roster and matches so you can upload a new file.",
                key="start_over_button",
            ):
                st.session_state.reset_confirm = True
                st.rerun()
        else:
            # Confirmation UI when reset_confirm is True
            st.warning(
                "Are you sure you want to **reset this meet**? "
                "This will clear the current roster, matchups, mat orders, exports, and undo history "
                "for this browser session."
            )
            c1, c2 = st.columns(2)
            with c1:
                if st.button("✅ Yes, reset meet", key="confirm_reset_yes"):
                    for key in [
                        "initialized", "bout_list", "mat_schedules", "suggestions",
                        "active", "mat_order", "excel_bytes", "pdf_bytes",
                        "roster", "manual_match_warning", "action_history"
                    ]:
                        st.session_state.pop(key, None)

                    # Reset confirmation flag
                    st.session_state.reset_confirm = False

                    # Bump uploader versions so Streamlit creates fresh, empty uploaders
                    st.session_state.roster_uploader_version += 1
                    st.session_state.state_json_uploader_version += 1  # clears JSON file selection

                    st.success("Meet reset. You can upload a new roster file.")
                    st.rerun()

            with c2:
                if st.button("❌ Cancel", key="confirm_reset_no"):
                    st.session_state.reset_confirm = False
                    st.info("Reset cancelled.")
                    st.rerun()

    # ----- Save / Load Meet (JSON snapshot) -----
    st.markdown("##### Save / Load Meet")

    # Export current meet to JSON
    if st.session_state.get("initialized"):
        snapshot = build_meet_snapshot()
        json_bytes = json.dumps(snapshot, indent=2).encode("utf-8")

        st.download_button(
            "💾 Download meet as JSON",
            data=json_bytes,
            file_name="wrestling_meet_state.json",
            mime="application/json",
            use_container_width=False,
        )

    # Import meet from JSON (manual load – avoids infinite restore loop)
    uploaded_state = st.file_uploader(
        "📂 Load saved meet (.json)",
        type="json",
        key=f"state_json_uploader_v{st.session_state.state_json_uploader_version}",
    )

    if uploaded_state is not None:
        if st.button("Load this saved meet", key="load_state_button"):
            try:
                data = json.load(uploaded_state)
                restore_meet_from_snapshot(data)
                st.success("Meet restored from JSON.")
                st.rerun()
            except Exception as e:
                st.error(f"Could not load saved meet: {e}")

    # Restore from server-side autosave file (if present)
    if os.path.exists(AUTOSAVE_FILE):
        if st.button("⏮️ Restore from autosave", key="restore_autosave_button"):
            try:
                with open(AUTOSAVE_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                restore_meet_from_snapshot(data)
                st.success("Meet restored from autosave.")
                st.rerun()
            except Exception as e:
                st.error(f"Could not restore autosave: {e}")

    # ----- Merge multiple roster CSV files (ALWAYS AVAILABLE) -----
    st.markdown("##### Merge multiple roster CSV files")
    st.caption(
        "Upload separate team roster CSV files and merge them into a single combined roster CSV. "
        "This does not change the current meet; it just helps you avoid manual copy/paste when "
        "building a master roster."
    )

    merge_files = st.file_uploader(
        "Select one or more roster CSV files",
        type=["csv"],
        accept_multiple_files=True,
        key="merge_rosters_simple",
    )

    # Helpful tip for coaches
    st.caption(
        "💡 *Tip:* After downloading a merged roster, you only need to come back here "
        "if a coach sends an updated CSV. Add or remove a file, then click Merge again "
        "to generate a new combined roster."
    )

    if merge_files:
        st.write("Files selected:")
        for f in merge_files:
            st.write("•", f.name)

        if st.button("Merge selected roster files", key="merge_rosters_button"):
            try:
                EXPECTED_COLUMNS = [
                    "name",
                    "team",
                    "grade",
                    "level",
                    "weight",
                    "early_matches",      # canonical name in your app
                    "scratch",
                    "gender",
                    "cross_gender_ok",
                ]

                dfs = []
                for f in merge_files:
                    df = pd.read_csv(f)

                    # Normalize column names
                    df.columns = [c.strip() for c in df.columns]

                    # Handle early_match vs early_matc typo
                    if "early_matc" in df.columns and "early_match" not in df.columns:
                        df = df.rename(columns={"early_matc": "early_match"})

                    # Ensure all expected columns exist
                    for col in EXPECTED_COLUMNS:
                        if col not in df.columns:
                            df[col] = None

                    # Reorder to your standard structure
                    df = df[EXPECTED_COLUMNS]
                    dfs.append(df)

                if not dfs:
                    st.error("No valid data found in the uploaded files.")
                else:
                    merged = pd.concat(dfs, ignore_index=True)

                    # Remove exact duplicates
                    merged = merged.drop_duplicates()

                    # Remove duplicate wrestler entries (same kid listed twice)
                    merged = merged.drop_duplicates(subset=["name", "team", "grade"])

                    st.success("✅ Rosters merged successfully.")

                    with st.expander("Show full merged roster (optional review)", expanded=False):
                        st.dataframe(merged, use_container_width=True)

                    csv_bytes = merged.to_csv(index=False).encode("utf-8-sig")
                    st.download_button(
                        label="Download merged_roster.csv",
                        data=csv_bytes,
                        file_name="merged_roster.csv",
                        mime="text/csv",
                        key="download_merged_roster",
                    )

                    # If in the future you want to auto-load this into the app, you could do:
                    # st.session_state["roster"] = merged.to_dict(orient="records")
                    # st.rerun()

            except Exception as e:
                st.error(f"Error during merge: {e}")

st.markdown("---")

# ----------------------------------------------------------------------
# SIDEBAR SETTINGS
# ----------------------------------------------------------------------
st.sidebar.header("Meet Settings")
st.sidebar.subheader("Search Wrestlers")
search_term = st.sidebar.text_input(
    "Filter by name or team",
    value="",
    placeholder="e.g. Smith or Red",
    key="wrestler_search",
    help="Search affects Mat Previews only (edit disabled while searching)."
)
st.sidebar.caption(
    "**Note:** Suggested Matches are based on all wrestlers; Mat Previews show only matches involving filtered wrestlers."
)

changed = False
st.sidebar.subheader("Match & Scheduling Rules")

# Top row: numbers in two columns
c1, c2 = st.sidebar.columns(2)
with c1:
    new_min = st.sidebar.number_input("Min Matches", 1, 10, CONFIG["MIN_MATCHES"], key="min_matches")
    new_max = st.sidebar.number_input("Max Matches", 1, 10, CONFIG["MAX_MATCHES"], key="max_matches")
    new_mats = st.sidebar.number_input("Number of Mats", 1, 10, CONFIG["NUM_MATS"], key="num_mats")
with c2:
    new_level_diff = st.sidebar.number_input("Max Level Diff", 0, 5, CONFIG["MAX_LEVEL_DIFF"], key="max_level_diff")
    new_min_weight = st.sidebar.number_input(
        "Min Wt Diff (lbs)", 0.0, 50.0, CONFIG["MIN_WEIGHT_DIFF"], 0.5,
        key="min_weight_diff"
    )

# Slider on its own row below the other settings
new_weight_factor = st.sidebar.slider(
    "Weight Diff % Factor",
    0.0, 0.5,
    CONFIG["WEIGHT_DIFF_FACTOR"],
    0.01,
    format="%.2f",
    key="weight_factor"
)

# Min rest gap
new_rest_gap = st.sidebar.number_input(
    "Min Rest Gap (matches)",
    1, 10,
    CONFIG.get("REST_GAP", 4),
    key="rest_gap"
)

if new_min > new_max:
    st.sidebar.error("Min Matches cannot exceed Max Matches!")
    new_min = new_max

st.sidebar.markdown("---")
st.sidebar.subheader("Team Colors")

circle_color_names = list(COLOR_ICON.keys())

# Rebuild TEAMS from the roster every run (if roster exists)
if st.session_state.get("roster"):
    roster_teams = sorted({
        str(w["team"]).strip()
        for w in st.session_state.roster
        if str(w["team"]).strip()
    })

    prev_teams = CONFIG.get("TEAMS", [])
    prev_color_by_name = {
        t["name"]: t["color"] for t in prev_teams if t.get("name")
    }

    TEAMS = []
    used_colors = set()

    for team_name in roster_teams:
        color = prev_color_by_name.get(team_name)
        if color not in circle_color_names:
            # pick first unused color, then wrap
            color = None
            for c in circle_color_names:
                if c not in used_colors:
                    color = c
                    break
            if color is None:
                color = circle_color_names[0]
        used_colors.add(color)
        TEAMS.append({"name": team_name, "color": color})

    CONFIG["TEAMS"] = TEAMS
    st.session_state.CONFIG = CONFIG
else:
    TEAMS = CONFIG.get("TEAMS", [])

if TEAMS:
    for i, team in enumerate(TEAMS):
        st.sidebar.markdown(f"**{team['name']}**")
        try:
            default_idx = circle_color_names.index(team["color"])
        except ValueError:
            default_idx = 0

        new_color = st.sidebar.selectbox(
            "Color",
            circle_color_names,
            index=default_idx,
            format_func=lambda x: x.capitalize(),
            key=f"color_{i}",
            label_visibility="collapsed"
        )

        if new_color != team["color"]:
            team["color"] = new_color
            changed = True
            st.session_state.sortable_version += 1
else:
    st.sidebar.caption("Upload a roster to configure team colors.")

if (
    new_min != CONFIG["MIN_MATCHES"] or new_max != CONFIG["MAX_MATCHES"] or
    new_mats != CONFIG["NUM_MATS"] or new_level_diff != CONFIG["MAX_LEVEL_DIFF"] or
    new_weight_factor != CONFIG["WEIGHT_DIFF_FACTOR"] or new_min_weight != CONFIG["MIN_WEIGHT_DIFF"] or
    new_rest_gap != CONFIG.get("REST_GAP", 4)
):
    CONFIG.update({
        "MIN_MATCHES": new_min,
        "MAX_MATCHES": new_max,
        "NUM_MATS": new_mats,
        "MAX_LEVEL_DIFF": new_level_diff,
        "WEIGHT_DIFF_FACTOR": new_weight_factor,
        "MIN_WEIGHT_DIFF": new_min_weight,
        "REST_GAP": new_rest_gap,
    })
    changed = True

st.sidebar.markdown("---")
if st.sidebar.button("Reset Settings", type="secondary"):
    # Reset CONFIG to BASE_CONFIG for this browser session only
    st.session_state.CONFIG = copy.deepcopy(BASE_CONFIG)
    CONFIG = st.session_state.CONFIG
    st.sidebar.success("Reset settings for this session. Refresh to apply.")
    st.rerun()

if changed:
    st.sidebar.success("Settings updated for this session. Refresh to apply.")
    st.rerun()

TEAM_COLORS = {t["name"]: COLOR_MAP.get(t["color"], "#000000") for t in TEAMS if t["name"]}
TEAM_COLOR_NAMES = {t["name"]: t["color"] for t in TEAMS if t["name"]}

# ----------------------------------------------------------------------
# MAIN APP – TABS
# ----------------------------------------------------------------------
if st.session_state.initialized:
    raw_active = st.session_state.active
    roster = st.session_state.roster

    tab_build, tab_summary, tab_help = st.tabs(["Match Builder", "Meet Summary", "Help"])

    # ==========================================================
    # TAB 1 – MATCH BUILDER
    # ==========================================================
    with tab_build:
        # Map each roster team to a color name (for icons + HTML)
        roster_teams = sorted({w["team"] for w in roster})
        palette = list(COLOR_ICON.keys())
        team_color_for_roster = {}

        # First, use explicit config team colors (TEAM_COLOR_NAMES)
        for team_name in roster_teams:
            cfg_color = TEAM_COLOR_NAMES.get(team_name)
            if cfg_color:
                team_color_for_roster[team_name] = cfg_color

        used_colors = set(team_color_for_roster.values())
        idx = 0
        for team_name in roster_teams:
            if team_name in team_color_for_roster:
                continue
            while palette[idx % len(palette)] in used_colors and len(used_colors) < len(palette):
                idx += 1
            color_name = palette[idx % len(palette)]
            team_color_for_roster[team_name] = color_name
            used_colors.add(color_name)
            idx += 1

        # ---------- QUICK START HELP (in-context) ----------
        st.markdown("### ❓ Quick Start")
        with st.expander("Quick Start Guide", expanded=False):
            st.markdown(
                """
        1. **Import your roster**
           - Use **Step 1** to download the roster template.
           - If you have **multiple team CSVs**, you can optionally use  
             **Advanced options → Merge multiple roster CSV files** to create a single combined `merged_roster.csv`.
           - In **Step 2**, upload your completed `roster.csv` (or `merged_roster.csv` if you merged files).
           - The app will auto-generate initial matchups once the file is uploaded.
        
        2. **Adjust meet settings (left sidebar)**
           - Set **Min / Max Matches**, **Number of Mats**, **Max Level Diff**, and **Min Wt Diff**.
           - Set **Min Rest Gap** so wrestlers don’t wrestle back-to-back.
           - After a roster is loaded, assign **team colors** (used in legends, emojis, Excel, and PDFs).
        
        3. **Apply scratches (before the meet)**
           - In **Pre-Meet Scratches**, select wrestlers who are not wrestling tonight (if they aren’t already flagged in the roster CSV).
           - Click **Apply scratches & regenerate schedule**. Early in the workflow (before manual edits), this will rebuild all matchups. After you’ve done manual editing, it will only remove matches involving scratched wrestlers and keep your mat layout.
           - Use **Start Over** if you want to completely rebuild from a fresh roster.
        
        4. **Fine-tune matchups**
           - Use **Manual Match Creator** to fill gaps for wrestlers under the minimum and when coaches want specific pairings.
           - In **Mat Previews**, drag rows to change bout order and remove individual bouts if needed.
        
        5. **Generate & download**
           - Click **Generate Coach Packets PDF** to build the Coach Packets that only contain team matches.
             - Use **Download Coach Packets PDF** to download document.
           - Click **Generate Documents** to build the **Excel Master Document** and **PDF Mat Printouts**.
             - Use **Download Excel** and **Download PDF** to download documents.
                """
            )

        # ----- Pre-Meet Scratches -----
        st.subheader("Pre-Meet Scratches")

        if roster:
            default_scratched_ids = [w["id"] for w in roster if w.get("scratch")]

            selected_scratched = st.multiselect(
                "Mark wrestlers as scratched (removed from meet scheduling):",
                options=[w["id"] for w in roster],
                default=default_scratched_ids,
                format_func=lambda wid: next(
                    f"{w['name']} ({w['team']})"
                    for w in roster if w["id"] == wid
                ),
            )

            apply_clicked = st.button("Apply scratches & regenerate schedule")

            st.caption(
                "Tip: After manual editing, applying scratches will only remove matches involving scratched wrestlers "
                "and keep your mat layout. Use **Start Over** if you want to completely rebuild all matchups."
            )

            if apply_clicked:
                # Take snapshot for undo **before** applying new scratches
                pre_snapshot = {
                    "roster": copy.deepcopy(st.session_state.roster),
                    "active": copy.deepcopy(st.session_state.active),
                    "bout_list": copy.deepcopy(st.session_state.bout_list),
                    "suggestions": copy.deepcopy(st.session_state.suggestions),
                    "mat_order": copy.deepcopy(st.session_state.mat_order),
                    "mat_overrides": copy.deepcopy(st.session_state.get("mat_overrides", {})),
                }

                # Update scratch flags based on selection
                for w in roster:
                    w["scratch"] = (w["id"] in selected_scratched)

                st.session_state.roster = roster
                new_active = [w for w in roster if not w["scratch"]]
                st.session_state.active = new_active

                existing_bouts = st.session_state.bout_list or []

                # Detect whether the meet is still in a "pristine" auto-generated state
                has_manual = any(b.get("manual") for b in existing_bouts)
                has_history = bool(st.session_state.get("action_history"))
                has_mat_order = any(st.session_state.mat_order.values())

                pristine = (not existing_bouts) or (not has_manual and not has_history and not has_mat_order)

                if pristine:
                    # Early workflow: behave like old logic – full regenerate
                    for w in roster:
                        w["match_ids"] = []
                    st.session_state.bout_list = generate_initial_matchups(new_active)
                    st.session_state.suggestions = build_suggestions(new_active, st.session_state.bout_list)
                    st.session_state.mat_order = {}
                    st.session_state.mat_overrides = {}
                    st.session_state.excel_bytes = None
                    st.session_state.pdf_bytes = None
                    st.session_state.action_history = []
                    st.session_state.sortable_version += 1

                    st.success("Scratches applied and schedule regenerated.")
                    st.rerun()
                else:
                    # Edited workflow: only remove matches

                    # Use snapshot captured BEFORE scratches were applied
                    push_action({
                        "type": "scratch_update",
                        "snapshot": pre_snapshot,
                    })

                    scratched_ids = {w["id"] for w in roster if w["scratch"]}

                    # Keep bouts that do NOT involve scratched wrestlers
                    remaining_bouts = [
                        b for b in existing_bouts
                        if b["w1_id"] not in scratched_ids and b["w2_id"] not in scratched_ids
                    ]

                    # Rebuild match_ids based on remaining bouts
                    for w in roster:
                        w["match_ids"] = []

                    for b in remaining_bouts:
                        w1 = next(w for w in roster if w["id"] == b["w1_id"])
                        w2 = next(w for w in roster if w["id"] == b["w2_id"])
                        w1["match_ids"].append(w2["id"])
                        w2["match_ids"].append(w1["id"])

                    st.session_state.bout_list = remaining_bouts

                    # Clean mat_order and mat_overrides to drop removed bouts
                    remaining_bout_nums = {b["bout_num"] for b in remaining_bouts}

                    cleaned_mat_order = {}
                    for mat, order in st.session_state.mat_order.items():
                        cleaned_order = [bn for bn in order if bn in remaining_bout_nums]
                        if cleaned_order:
                            cleaned_mat_order[mat] = cleaned_order
                    st.session_state.mat_order = cleaned_mat_order

                    overrides = st.session_state.get("mat_overrides", {})
                    st.session_state.mat_overrides = {
                        bn: m for bn, m in overrides.items() if bn in remaining_bout_nums
                    }

                    # Rebuild suggestions based on new active + remaining bouts
                    st.session_state.suggestions = build_suggestions(new_active, remaining_bouts)

                    # Invalidate exports; refresh drag widgets
                    st.session_state.excel_bytes = None
                    st.session_state.pdf_bytes = None
                    st.session_state.sortable_version += 1

                    st.success(
                        "Scratches applied: matches involving scratched wrestlers were removed. "
                        "Manual matches and mat layout for remaining bouts were preserved."
                    )
                    st.rerun()

        # ---- Filtered wrestlers by search ----
        if search_term.strip():
            term = search_term.strip().lower()
            filtered_active = [
                w for w in raw_active
                if term in w["name"].lower() or term in w["team"].lower()
            ]
            st.info(
                f"Showing **{len(filtered_active)}** wrestler(s) matching “{search_term}” "
                f"(out of {len(raw_active)} active)."
            )
        else:
            filtered_active = raw_active
            st.info(f"Showing **all {len(filtered_active)}** active wrestlers.")

        filtered_ids = {w["id"] for w in filtered_active}

        # ----- Manual Match Creator -----
        st.subheader("Manual Match Creator")

        # Show any stored manual-match warning from last run
        manual_warning = st.session_state.get("manual_match_warning")
        if manual_warning:
            st.warning(manual_warning)
            st.session_state.manual_match_warning = ""

        active_ids = [w["id"] for w in raw_active]

        if len(active_ids) < 2:
            st.caption("Not enough active wrestlers to create a manual match.")
        else:
            # Map IDs to wrestler records for quick lookup
            id_to_wrestler = {w["id"]: w for w in raw_active}

            # NEW: helper for gender tags in manual match dropdowns
            def gender_tag_from_id(wid: int) -> str:
                w = id_to_wrestler.get(wid)
                if not w:
                    return "-"
                g = w.get("gender")
                return g if g in ("M", "F") else "-"

            # All active wrestlers sorted by weight (lightest → heaviest)
            sorted_all_ids = sorted(active_ids, key=lambda wid: id_to_wrestler[wid]["weight"])

            # ---- NEW: Wrestler 1 filter toggle ----
            w1_filter_mode = st.radio(
                "Wrestler 1 list",
                options=[
                    "Show everyone",
                    "Only wrestlers below MIN matches",
                ],
                horizontal=True,
                key="manual_w1_filter_mode",
                help=(
                    "Show either all active wrestlers, or only those who currently have fewer "
                    f"than MIN matches ({CONFIG['MIN_MATCHES']}). Wrestler 2 stays unfiltered."
                ),
            )

            if w1_filter_mode == "Only wrestlers below MIN matches":
                filtered_ids_for_w1 = [
                    wid for wid in sorted_all_ids
                    if len(id_to_wrestler[wid]["match_ids"]) < CONFIG["MIN_MATCHES"]
                ]
                # If everyone already meets the minimum, fall back to all wrestlers
                if not filtered_ids_for_w1:
                    st.info(
                        "All wrestlers already meet the minimum matches – "
                        "showing everyone for Wrestler 1."
                    )
                    filtered_ids_for_w1 = sorted_all_ids
            else:
                filtered_ids_for_w1 = sorted_all_ids

            # Percentage of roster to consider around Wrestler 1 (for Wrestler 2)
            # e.g. 0.30 = 30% of wrestlers centered around Wrestler 1's weight
            WINDOW_PCT = 0.30

            col_m1, col_m2 = st.columns([3, 3])

            # ---------------- Wrestler 1 ----------------
            with col_m1:
                manual_w1_id = st.selectbox(
                    "Wrestler 1",
                    options=filtered_ids_for_w1,
                    format_func=lambda wid: (
                        f"{id_to_wrestler[wid]['name']} "
                        f"({id_to_wrestler[wid]['team']}, {gender_tag_from_id(wid)}) – "
                        f"Lvl {id_to_wrestler[wid]['level']:.1f}, "
                        f"{id_to_wrestler[wid]['weight']:.0f} lbs, "
                        f"Matches: {len(id_to_wrestler[wid]['match_ids'])}"
                    ),
                    key="manual_match_w1",
                )

            # ---------------- Wrestler 2 ----------------
            # NOTE: Wrestler 2 stays based on the full list (can go over MAX matches)
            with col_m2:
                if manual_w1_id is not None and manual_w1_id in sorted_all_ids:
                    total = len(sorted_all_ids)
                    window_size = max(1, int(total * WINDOW_PCT))

                    # Index of Wrestler 1 in the global weight-sorted list
                    center_idx = sorted_all_ids.index(manual_w1_id)
                    half = window_size // 2
                    start = max(0, center_idx - half)
                    end = min(total, center_idx + half + 1)

                    # Wrestlers who already have a match with Wrestler 1
                    w1_existing_opponents = set(id_to_wrestler[manual_w1_id]["match_ids"])

                    # Filter: within window, not W1, not already matched with W1,
                    # and gender-compatible (Option A)
                    candidate_ids = [
                        wid for wid in sorted_all_ids[start:end]
                        if wid != manual_w1_id
                           and wid not in w1_existing_opponents
                           and genders_compatible(
                                id_to_wrestler[manual_w1_id],
                                id_to_wrestler[wid]
                            )
                    ]

                    # Fallback: if window collapses, use all others not already opponents,
                    # still respecting gender compatibility.
                    if not candidate_ids:
                        candidate_ids = [
                            wid for wid in sorted_all_ids
                            if wid != manual_w1_id
                               and wid not in w1_existing_opponents
                               and genders_compatible(
                                    id_to_wrestler[manual_w1_id],
                                    id_to_wrestler[wid]
                                )
                        ]
                else:
                    w1_existing_opponents = set(
                        id_to_wrestler.get(manual_w1_id, {}).get("match_ids", [])
                    )
                    candidate_ids = [
                        wid for wid in sorted_all_ids
                        if wid != manual_w1_id
                           and wid not in w1_existing_opponents
                           and genders_compatible(
                                id_to_wrestler[manual_w1_id],
                                id_to_wrestler[wid]
                            )
                    ]

                manual_w2_id = st.selectbox(
                    "Wrestler 2",
                    options=candidate_ids,
                    format_func=lambda wid: (
                        f"{id_to_wrestler[wid]['name']} "
                        f"({id_to_wrestler[wid]['team']}, {gender_tag_from_id(wid)}) – "
                        f"Lvl {id_to_wrestler[wid]['level']:.1f}, "
                        f"{id_to_wrestler[wid]['weight']:.0f} lbs, "
                        f"Matches: {len(id_to_wrestler[wid]['match_ids'])}"
                    ),
                    key="manual_match_w2",
                )

                # nest a small two-column layout just for right-aligning the button
                btn_spacer, btn_col = st.columns([3, 1])
                with btn_col:
                    create_manual = st.button(
                        "Create Match",
                        use_container_width=True,
                        help="Force a match between these two wrestlers, even if it wasn’t auto-generated.",
                        key="manual_match_create_btn",
                    )

            if create_manual:
                if manual_w1_id == manual_w2_id:
                    st.warning("Please choose two different wrestlers.")
                else:
                    w1 = next(w for w in raw_active if w["id"] == manual_w1_id)
                    w2 = next(w for w in raw_active if w["id"] == manual_w2_id)

                    # Extra safety: don't allow gender-incompatible pair, even though
                    # we filtered them out of the dropdown.
                    if not genders_compatible(w1, w2):
                        st.warning(
                            "This manual pairing does not respect gender preferences and cannot be created. "
                            "Adjust the wrestlers' gender or cross-gender settings if this match is intended."
                        )
                        st.stop()

                    # Check if they already have a match together (ignore Manually Removed bouts)
                    already_linked = any(
                        (
                            (b["w1_id"] == w1["id"] and b["w2_id"] == w2["id"]) or
                            (b["w1_id"] == w2["id"] and b["w2_id"] == w1["id"])
                        )
                        for b in st.session_state.bout_list
                        if b.get("manual") != "Manually Removed"
                    )

                    if already_linked:
                        msg = (
                            f"{w1['name']} ({w1['team']}) and "
                            f"{w2['name']} ({w2['team']}) already have a match together. "
                            "A new match will not be created."
                        )
                        st.session_state.manual_match_warning = msg
                        st.warning(msg)
                    else:
                        # Soft warnings for coaches – but still allow the match
                        warning_msgs = []
                        if w1["team"] == w2["team"]:
                            warning_msgs.append("Same team matchup.")
                        if abs(w1["level"] - w2["level"]) > CONFIG["MAX_LEVEL_DIFF"]:
                            warning_msgs.append("Large level difference.")
                        if abs(w1["weight"] - w2["weight"]) > max_weight_diff(w1["weight"]):
                            warning_msgs.append("Large weight difference.")

                        if warning_msgs:
                            st.info(
                                "Note: " + " ".join(
                                    f"• {msg}" for msg in warning_msgs
                                ) + " (match will still be created)."
                            )

                        # Link in match_ids if not already present
                        if w2["id"] not in w1["match_ids"]:
                            w1["match_ids"].append(w2["id"])
                        if w1["id"] not in w2["match_ids"]:
                            w2["match_ids"].append(w1["id"])

                        new_bout_num = (max([b["bout_num"] for b in st.session_state.bout_list]) + 1) \
                            if st.session_state.bout_list else 1

                        new_score = matchup_score(w1, w2)
                        new_bout = {
                            "bout_num": new_bout_num,
                            "w1_id": w1["id"], "w1_name": w1["name"], "w1_team": w1["team"],
                            "w1_level": w1["level"], "w1_weight": w1["weight"],
                            "w1_grade": w1["grade"], "w1_early": w1["early"],
                            "w2_id": w2["id"], "w2_name": w2["name"], "w2_team": w2["team"],
                            "w2_level": w2["level"], "w2_weight": w2["weight"],
                            "w2_grade": w2["grade"], "w2_early": w2["early"],
                            "score": new_score,
                            "avg_weight": (w1["weight"] + w2["weight"]) / 2,
                            "is_early": w1["early"] or w2["early"],
                            "manual": "Coach Manual Match",
                        }

                        st.session_state.bout_list.append(new_bout)

                        # Keep bouts sorted by avg_weight so base scheduler behaves
                        st.session_state.bout_list.sort(key=lambda x: x["avg_weight"])

                        # Clear manual mat order so the new match gets placed, then coach can drag it
                        st.session_state.mat_order = {}

                        # Rebuild suggestions based on new counts
                        st.session_state.suggestions = build_suggestions(raw_active, st.session_state.bout_list)

                        # Invalidate exports
                        st.session_state.excel_bytes = None
                        st.session_state.pdf_bytes = None

                        # Record action for undo
                        push_action({"type": "manual_add", "bout_num": new_bout_num})

                        # Refresh drag widgets
                        st.session_state.sortable_version += 1

                        st.success(
                            f"Manual match created: {w1['name']} ({w1['team']}) vs {w2['name']} ({w2['team']}). "
                            "You can now drag it to the desired mat and slot."
                        )
                        st.rerun()

        # ----- Global schedule & rest conflicts -----
        full_schedule = apply_mat_order_to_global_schedule() if st.session_state.bout_list else []
        rest_gap = CONFIG.get("REST_GAP", 4)
        conflicts_all = compute_rest_conflicts(full_schedule, rest_gap) if full_schedule else []

        # NEW: multi-mat warning
        multi_mat_issues = compute_multi_mat_assignments(full_schedule) if full_schedule else []
        multi_mat_ids = {issue["wrestler_id"] for issue in multi_mat_issues} if multi_mat_issues else set()
        
        if multi_mat_issues:
            st.warning(
                f"{len(multi_mat_issues)} wrestler(s) are assigned to matches on more than one mat."
            )
            with st.expander("Show wrestlers on multiple mats", expanded=False):
                for issue in multi_mat_issues:
                    # Build per-mat slot display like: "1 (Slot 3), 2 (Slots 4, 9)"
                    parts = []
                    for mat in issue["mats"]:
                        slots_on_mat = sorted(
                            m["slot"]
                            for m in issue["matches"]
                            if m["mat"] == mat
                        )
                        if len(slots_on_mat) == 1:
                            slot_text = f"Slot {slots_on_mat[0]}"
                        else:
                            slot_text = "Slots " + ", ".join(str(s) for s in slots_on_mat)
        
                        parts.append(f"{mat} ({slot_text})")
        
                    mat_slot_text = ", ".join(parts)
                    st.markdown(
                        f"- **{issue['name']}** ({issue['team']}): Mats {mat_slot_text}"
                    )
        else:
            st.caption("All wrestlers are currently assigned to a single mat.")
        

        if search_term.strip():
            visible_conflicts = [c for c in conflicts_all if c["wrestler_id"] in filtered_ids]
        else:
            visible_conflicts = conflicts_all

        st.subheader("Mat Previews")

        # NEW: map ID -> wrestler for gender display on mat previews
        id_to_wrestler_global = {w["id"]: w for w in roster}

        def gender_display(wid: int) -> str:
            w = id_to_wrestler_global.get(wid)
            if not w:
                return "?"
            g = w.get("gender")
            return g if g in ("M", "F") else "?"

        if visible_conflicts:
            st.warning(
                f"Rest conflicts detected: **{len(visible_conflicts)}** (requires at least "
                f"**{rest_gap}** matches between bouts for the same wrestler)."
            )
        else:
            st.caption(f"No rest conflicts found (min gap: {rest_gap} matches).")

        if not full_schedule:
            st.caption("No bouts scheduled yet.")
        else:
            def bout_in_filtered(b):
                return (
                    b["manual"] != "Manually Removed" and
                    (b["w1_id"] in filtered_ids or b["w2_id"] in filtered_ids)
                )

            # ---------- SEARCH MODE (read-only, HTML table) ----------
            if search_term.strip():
                for mat in range(1, CONFIG["NUM_MATS"] + 1):
                    mat_entries = [
                        e for e in full_schedule
                        if e["mat"] == mat and bout_in_filtered(
                            next(
                                b for b in st.session_state.bout_list
                                if b["bout_num"] == e["bout_num"]
                            )
                        )
                    ]
                    mat_label = f"Mat {mat} ({len(mat_entries)} matches)"
                    with st.expander(mat_label, expanded=True):
                        if not mat_entries:
                            st.caption("No matches for the current filter on this mat.")
                            continue

                        # HTML table with colored dots
                        table_rows = []
                        for e in mat_entries:
                            b = next(
                                x for x in st.session_state.bout_list
                                if x["bout_num"] == e["bout_num"]
                            )
                            early_flag = "⏰🔥 EARLY 🔥⏰" if b["is_early"] else ""
                            color_name1 = team_color_for_roster.get(b["w1_team"])
                            color_name2 = team_color_for_roster.get(b["w2_team"])
                            dot1 = color_dot_hex(COLOR_MAP.get(color_name1, "#000000")) if color_name1 else ""
                            dot2 = color_dot_hex(COLOR_MAP.get(color_name2, "#000000")) if color_name2 else ""
                            g1 = gender_display(b["w1_id"])
                            g2 = gender_display(b["w2_id"])

                            table_rows.append(
                            f"<tr>"
                            f"<td>{e['mat_bout_num']}</td>"
                            f"<td>{early_flag}</td>"
                            f"<td>{dot1}{b['w1_name']} ({b['w1_team']}, {g1})</td>"
                            f"<td>{dot2}{b['w2_name']} ({b['w2_team']}, {g2})</td>"
                            f"<td>{b['w1_level']:.1f}/{b['w2_level']:.1f}</td>"
                            f"<td>{b['w1_weight']:.0f}/{b['w2_weight']:.0f}</td>"
                            f"<td>{b['score']:.1f}</td>"
                            f"</tr>"
                        )
                        
                        table_html = (
                            "<table style='width:100%;border-collapse:collapse;font-size:0.80rem;'>"
                            "<thead>"
                            "<tr style='background:#f0f0f0;'>"
                            "<th style='border:1px solid #ddd;padding:4px;'>Slot</th>"
                            "<th style='border:1px solid #ddd;padding:4px;'>Early</th>"
                            "<th style='border:1px solid #ddd;padding:4px;'>Wrestler 1</th>"
                            "<th style='border:1px solid #ddd;padding:4px;'>Wrestler 2</th>"
                            "<th style='border:1px solid #ddd;padding:4px;'>Lvls</th>"
                            "<th style='border:1px solid #ddd;padding:4px;'>Wts</th>"
                            "<th style='border:1px solid #ddd;padding:4px;'>Score</th>"
                            "</tr>"
                            "</thead>"
                            "<tbody>"
                            + "".join(table_rows) +
                            "</tbody>"
                            "</table>"
                        )

                        st.markdown(table_html, unsafe_allow_html=True)

                        # Per-mat rest warnings for visible wrestlers
                        mat_conflicts = [
                            c for c in visible_conflicts if c["mat"] == mat
                        ]
                        if mat_conflicts:
                            st.markdown("**Rest warnings on this mat (filtered wrestlers):**")
                            for c in mat_conflicts:
                                st.markdown(
                                    f"- {c['wrestler']} ({c['team']}): "
                                    f"Slot {c['slot1']} → Slot {c['slot2']} "
                                    f"(gap {c['gap']} < required {rest_gap})"
                                )

                st.caption("Reordering and removal are disabled while search is active. Clear the search box to edit mats.")

            # ---------- EDIT MODE (drag + per-mat remove + move) ----------
            else:
                for mat in range(1, CONFIG["NUM_MATS"] + 1):
                    mat_entries = [e for e in full_schedule if e["mat"] == mat]
                    mat_label = f"Mat {mat} ({len(mat_entries)} matches)"
                    with st.expander(mat_label, expanded=True):
                        if not mat_entries:
                            st.caption("No bouts on this mat.")
                            continue

                        bout_nums_in_mat = [e["bout_num"] for e in mat_entries]
                        existing_order = st.session_state.mat_order.get(mat)
                        if not existing_order:
                            st.session_state.mat_order[mat] = bout_nums_in_mat.copy()
                        else:
                            cleaned = [bn for bn in existing_order if bn in bout_nums_in_mat]
                            for bn in bout_nums_in_mat:
                                if bn not in cleaned:
                                    cleaned.append(bn)
                            st.session_state.mat_order[mat] = cleaned

                        prev_order = st.session_state.mat_order[mat].copy()

                        # Legend for teams on this mat (HTML dots)
                        teams_on_mat = set()
                        for e in mat_entries:
                            b_for_legend = next(
                                x for x in st.session_state.bout_list
                                if x["bout_num"] == e["bout_num"]
                            )
                            teams_on_mat.add(b_for_legend["w1_team"])
                            teams_on_mat.add(b_for_legend["w2_team"])
                        legend_bits = []
                        for t in sorted(teams_on_mat):
                            hex_color = TEAM_COLORS.get(t, "#000000")
                            dot = color_dot_hex(hex_color)
                            legend_bits.append(f"{dot}{t}")
                        if legend_bits:
                            legend_html = " ".join(legend_bits)
                            st.markdown(
                                f"<div style='margin-bottom:4px;font-size:0.8rem;'>Teams on this mat: {legend_html}</div>",
                                unsafe_allow_html=True,
                            )

                        # Build drag labels (plain text, circle emojis + gender)
                        row_labels = []
                        label_to_bout = {}
                        for slot_index, bn in enumerate(st.session_state.mat_order[mat], start=1):
                            if bn not in bout_nums_in_mat:
                                continue
                            b = next(x for x in st.session_state.bout_list if x["bout_num"] == bn)

                            early_prefix = "🔥🔥⏰ EARLY MATCH ⏰🔥🔥  |  " if b["is_early"] else ""

                            color_name1 = team_color_for_roster.get(b["w1_team"])
                            color_name2 = team_color_for_roster.get(b["w2_team"])
                            icon1 = COLOR_ICON.get(color_name1, "●")
                            icon2 = COLOR_ICON.get(color_name2, "●")
                            g1 = gender_display(b["w1_id"])
                            g2 = gender_display(b["w2_id"])

                            label = (
                            f"{early_prefix}"
                            f"Slot {slot_index:02d} | "
                            f"{icon1} {b['w1_name']} ({b['w1_team']}, {g1})  vs  "
                            f"{icon2} {b['w2_name']} ({b['w2_team']}, {g2})"
                            f"  |  Lvl {b['w1_level']:.1f}/{b['w2_level']:.1f}"
                            f"  |  Wt {b['w1_weight']:.0f}/{b['w2_weight']:.0f}"
                            f"  |  Score {b['score']:.1f}"
                        )

                            row_labels.append(label)
                            label_to_bout[label] = bn

                        sorted_labels = sort_items(
                            row_labels,
                            direction="vertical",
                            key=f"mat_{mat}_sortable_v{st.session_state.sortable_version}",
                            custom_style=SORTABLE_STYLE,
                        )

                        new_order = []
                        for label in sorted_labels:
                            bn = label_to_bout.get(label)
                            if bn is not None and bn in bout_nums_in_mat and bn not in new_order:
                                new_order.append(bn)

                        if new_order != prev_order:
                            # Take a snapshot of current mat_order for unified undo
                            snapshot_order = {
                                m: order.copy() for m, order in st.session_state.mat_order.items()
                            }
                            push_action({
                                "type": "drag",
                                "previous_mat_order": snapshot_order,
                            })

                            st.session_state.mat_order[mat] = new_order
                            st.session_state.excel_bytes = None
                            st.session_state.pdf_bytes = None
                            st.session_state.sortable_version += 1
                            st.rerun()
                        else:
                            st.session_state.mat_order[mat] = new_order

                        st.caption("Drag rows above – top row is Slot 1, next is Slot 2, etc. for this mat.")

                        # Per-mat remove + move
                        bout_label_map = {}
                        for idx2, bn in enumerate(st.session_state.mat_order[mat], start=1):
                            if bn not in bout_nums_in_mat:
                                continue
                            b = next(x for x in st.session_state.bout_list if x["bout_num"] == bn)
                            bout_label_map[bn] = (
                                f"Slot {idx2}: "
                                f"{b['w1_name']} ({b['w1_team']}) vs {b['w2_name']} ({b['w2_team']})"
                            )


                        valid_bouts = list(bout_label_map.keys())
                        if not valid_bouts:
                            st.caption("No bouts left on this mat.")
                        else:
                            selected_bout = st.selectbox(
                                "Remove bout on this mat:",
                                options=valid_bouts,
                                format_func=lambda v: bout_label_map[v],
                                key=f"remove_select_mat_{mat}"
                            )
                            if st.button(
                                "Remove selected bout",
                                key=f"remove_button_mat_{mat}",
                                help="Removes the selected bout from this meet (Undo available below)."
                            ):
                                remove_bout(selected_bout)

                            # --- Move selected bout to another mat (button below selectbox) ---
                            st.markdown("**Move selected bout to another mat**")

                            move_target_mat = st.selectbox(
                                "Target mat",
                                options=[m for m in range(1, CONFIG["NUM_MATS"] + 1) if m != mat],
                                key=f"move_target_mat_{mat}",
                            )

                            # Create a full-width container to match the Remove button sizing
                            move_button_area = st.container()
                            with move_button_area:
                                if st.button(
                                    "Move to mat",
                                    key=f"move_button_mat_{mat}",
                                    help="Move the selected bout to the chosen mat."
                                ):
                                    # Update mat_overrides so the scheduler keeps it on the new mat
                                    overrides = st.session_state.get("mat_overrides", {})
                                    overrides[selected_bout] = move_target_mat
                                    st.session_state.mat_overrides = overrides

                                    # Update mat_order: remove from this mat, append to target mat
                                    src_order = st.session_state.mat_order.get(mat, [])
                                    if selected_bout in src_order:
                                        src_order.remove(selected_bout)
                                    st.session_state.mat_order[mat] = src_order

                                    dest_order = st.session_state.mat_order.get(move_target_mat, [])
                                    if selected_bout not in dest_order:
                                        dest_order.append(selected_bout)
                                    st.session_state.mat_order[move_target_mat] = dest_order

                                    # Invalidate exports and refresh drag widgets
                                    st.session_state.excel_bytes = None
                                    st.session_state.pdf_bytes = None
                                    st.session_state.sortable_version += 1

                                    st.success(
                                        f"Bout {selected_bout} moved to Mat {move_target_mat}. "
                                        "You can now reorder it on that mat."
                                    )
                                    st.rerun()

                        # Per-mat rest warnings (all wrestlers)
                        mat_conflicts = [c for c in visible_conflicts if c["mat"] == mat]
                        if mat_conflicts:
                            lines = []
                            for c in mat_conflicts:
                                lines.append(
                                    f"- {c['wrestler']} ({c['team']}): "
                                    f"Slot {c['slot1']} → Slot {c['slot2']} "
                                    f"(gap {c['gap']} < required {rest_gap})"
                                )
                        
                            st.warning(
                                "**Rest warnings on this mat:**\n" + "\n".join(lines)
                            )



        # ----- Unified Undo Button -----
        st.markdown("---")

        last_action = st.session_state.action_history[-1] if st.session_state.action_history else None

        if last_action:
            t = last_action.get("type")
            if t == "remove":
                label = "Undo Last Remove"
            elif t == "drag":
                label = "Undo Last Drag / Reorder"
            elif t == "manual_add":
                label = "Undo Last Manual Match"
            elif t == "suggest_add":
                label = "Undo Last Suggested Matches"
            elif t == "scratch_update":
                label = "Undo Last Scratches Update"
            else:
                label = "Undo Last Action"

            if st.button(label, help="Undo the most recent change (remove/drag/manual/suggested/scratches)"):
                undo_last_action()
        else:
            st.caption("No actions yet to undo.")

        # ================================
        # ---- COACH PACKETS (PER TEAM) ---
        # ================================
        st.markdown("---")
        st.markdown("### Coach Packets (per team)")

        generate_coach = st.button(
            "Generate Coach Packets PDF",
            type="primary",  # red button
            help="Builds a page per team with all matches for each wrestler.",
            key="generate_coach_packets_btn",
        )

        if generate_coach:
            if not full_schedule:
                st.warning("No schedule yet – build matchups first.")
            else:
                try:
                    coach_pdf = generate_coach_packets_pdf(full_schedule)
                    st.session_state.coach_pdf_bytes = coach_pdf
                    st.toast("Coach packets PDF generated.", icon="📄")
                except Exception as e:
                    st.error(f"Could not generate coach packets: {e}")

        # Download button for coach packets
        if st.session_state.get("coach_pdf_bytes"):
            st.download_button(
                "Download Coach Packets PDF",
                data=st.session_state.coach_pdf_bytes,
                file_name="coach_packets.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        # ================================
        # ---- MEET DOCUMENTS SECTION ----
        # ================================
        st.markdown("---")
        st.markdown("### Meet Documents (Excel + Mat PDF)")

        # ---- GENERATE MEET (Excel + Mat PDFs) ----
        if st.button(
            "Generate Documents",
            type="primary",
            help="Generate Excel + mat-by-mat PDF for download",
            key="generate_meet_docs_btn",
        ):
            with st.spinner("Generating files..."):
                try:
                    final_sched = apply_mat_order_to_global_schedule()
                    st.session_state.mat_schedules = final_sched

                    # -------- Excel generation --------
                    out = io.BytesIO()
                    with pd.ExcelWriter(out, engine="openpyxl") as writer:
                        # Roster sheet (active wrestlers only)
                        roster_df = pd.DataFrame(st.session_state.active)
                        roster_df.to_excel(writer, sheet_name="Roster", index=False)

                        # All matchups sheet
                        matchups_df = pd.DataFrame(st.session_state.bout_list)
                        matchups_df.to_excel(writer, sheet_name="Matchups", index=False)

                        # Remaining suggestions
                        suggestions_df = pd.DataFrame(st.session_state.suggestions)
                        suggestions_df.to_excel(writer, sheet_name="Remaining Suggestions", index=False)

                        # Per-mat sheets
                        for m in range(1, CONFIG["NUM_MATS"] + 1):
                            data = [e for e in final_sched if e["mat"] == m]
                            if not data:
                                pd.DataFrame(
                                    [["", "", ""]],
                                    columns=["#", "Wrestler 1 (Team)", "Wrestler 2 (Team)"],
                                ).to_excel(writer, f"Mat {m}", index=False)
                                continue

                            df = pd.DataFrame(data)[["mat_bout_num", "w1", "w2"]]
                            df.columns = ["#", "Wrestler 1 (Team)", "Wrestler 2 (Team)"]
                            df.to_excel(writer, f"Mat {m}", index=False)

                            # Highlight early matches if openpyxl is available
                            if _EXCEL_AVAILABLE:
                                ws = writer.book[f"Mat {m}"]
                                fill = PatternFill(
                                    start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                                )
                                for i, _ in df.iterrows():
                                    if next(
                                        b
                                        for b in st.session_state.bout_list
                                        if b["bout_num"] == data[i]["bout_num"]
                                    )["is_early"]:
                                        for c in range(1, 3 + 1):
                                            ws.cell(row=i + 2, column=c).fill = fill

                    st.session_state.excel_bytes = out.getvalue()

                    # -------- Mat-by-mat PDF generation --------
                    buf = io.BytesIO()
                    doc = SimpleDocTemplate(buf, pagesize=letter)
                    elements = []
                    styles = getSampleStyleSheet()

                    for m in range(1, CONFIG["NUM_MATS"] + 1):
                        data = [e for e in final_sched if e["mat"] == m]
                        if not data:
                            elements.append(Paragraph(f"Mat {m} - No matches", styles["Title"]))
                            elements.append(PageBreak())
                            continue

                        table = [["#", "Wrestler 1", "Wrestler 2"]]
                        for e in data:
                            b = next(
                                x
                                for x in st.session_state.bout_list
                                if x["bout_num"] == e["bout_num"]
                            )
                            table.append(
                                [
                                    e["mat_bout_num"],
                                    Paragraph(
                                        f'<font color="{TEAM_COLORS.get(b["w1_team"], "#000")}">'
                                        f"<b>{b['w1_name']}</b></font> ({b['w1_team']})",
                                        styles["Normal"],
                                    ),
                                    Paragraph(
                                        f'<font color="{TEAM_COLORS.get(b["w2_team"], "#000")}">'
                                        f"<b>{b['w2_name']}</b></font> ({b['w2_team']})",
                                        styles["Normal"],
                                    ),
                                ]
                            )

                        t = Table(table, colWidths=[0.5 * inch, 3 * inch, 3 * inch])
                        s = TableStyle(
                            [
                                ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.black),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.lightgrey),
                                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ]
                        )

                        # Highlight early matches
                        for r, _ in enumerate(table[1:], 1):
                            if next(
                                b
                                for b in st.session_state.bout_list
                                if b["bout_num"] == data[r - 1]["bout_num"]
                            )["is_early"]:
                                s.add("BACKGROUND", (0, r), (-1, r), HexColor("#FFFF99"))

                        t.setStyle(s)
                        elements += [Paragraph(f"Mat {m}", styles["Title"]), Spacer(1, 12), t]
                        if m < CONFIG["NUM_MATS"]:
                            elements.append(PageBreak())

                    doc.build(elements)
                    st.session_state.pdf_bytes = buf.getvalue()

                    st.toast("Meet documents generated!", icon="✅")
                except Exception as e:
                    st.error(f"Generation failed: {e}")
                    st.toast("Error – check console.", icon="⚠️")

        # Download buttons for meet documents
        col_ex, col_pdf = st.columns(2)
        with col_ex:
            if st.session_state.excel_bytes is not None:
                st.download_button(
                    label="Download Excel",
                    data=st.session_state.excel_bytes,
                    file_name="meet_schedule.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        with col_pdf:
            if st.session_state.pdf_bytes is not None:
                st.download_button(
                    label="Download Mat PDF",
                    data=st.session_state.pdf_bytes,
                    file_name="meet_schedule.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
    # ==========================================================
    # TAB 2 – MEET SUMMARY
    # ==========================================================
    with tab_summary:
        st.subheader("Meet Summary")

        full_schedule = apply_mat_order_to_global_schedule() if st.session_state.bout_list else []
        rest_gap = CONFIG.get("REST_GAP", 4)
        conflicts_all = compute_rest_conflicts(full_schedule, rest_gap) if full_schedule else []

        num_wrestlers = len(st.session_state.active)
        total_bouts = len([b for b in st.session_state.bout_list if b["manual"] != "Manually Removed"])
        avg_matches = (
            total_bouts * 2 / num_wrestlers if num_wrestlers > 0 else 0.0
        )

        c1, c2, c3 = st.columns(3)
        c1.metric("Active Wrestlers", num_wrestlers)
        num_scratched = len([w for w in st.session_state.roster if w.get("scratch")])
        c1.metric("Scratched Wrestlers", num_scratched)
        c2.metric("Total Bouts", total_bouts)
        c3.metric("Avg Matches / Wrestler", f"{avg_matches:.2f}")

        st.markdown("---")

        # Wrestler Match Counts with Grade, Level, Weight + proper sorting
        st.markdown("#### Wrestler Match Counts")

        valid_bouts = [b for b in st.session_state.bout_list if b["manual"] != "Manually Removed"]
        if not st.session_state.active:
            st.caption("No wrestlers yet.")
        else:
            # Build match counts
            match_counts = {}
            for b in valid_bouts:
                for side in ("w1", "w2"):
                    wid = b[f"{side}_id"]
                    if wid not in match_counts:
                        match_counts[wid] = {"Matches": 0}
                    match_counts[wid]["Matches"] += 1

            # Build rows with full wrestler data, including gender
            rows = []
            for w in st.session_state.active:
                rows.append({
                    "Wrestler": w["name"],
                    "Team": w["team"],
                    "Grade": w["grade"],
                    "Level": f"{w['level']:.1f}",
                    "Weight": w["weight"],           # keep as float for correct sorting
                    "Weight_display": f"{w['weight']:.0f}",  # nice display version
                    "Gender": (w.get("gender") if w.get("gender") in ("M", "F") else "Unknown"),
                    "Matches": match_counts.get(w["id"], {}).get("Matches", 0),
                })

            df_wc = pd.DataFrame(rows)

            # Status column
            min_m = CONFIG["MIN_MATCHES"]
            max_m = CONFIG["MAX_MATCHES"]
            df_wc["Status"] = df_wc["Matches"].apply(
                lambda m: "Below Min" if m < min_m else ("Above Max" if m > max_m else "OK")
            )

            # NEW: gender filter
            gender_options = ["M", "F", "Unknown"]
            selected_genders = st.multiselect(
                "Filter by gender",
                options=gender_options,
                default=gender_options,
                key="summary_gender_filter"
            )
            df_wc = df_wc[df_wc["Gender"].isin(selected_genders)]

            # Default sort: Team → Wrestler name
            default_df = df_wc.sort_values(["Team", "Wrestler"]).reset_index(drop=True)

            # Add a sort selector
            sort_by = st.radio(
                "Sort table by:",
                options=["Team (default)", "Weight (light → heavy)"],
                horizontal=True,
                index=0,
                key="summary_sort"
            )

            if sort_by == "Weight (light → heavy)":
                display_df = df_wc.sort_values("Weight").reset_index(drop=True)
            else:
                display_df = default_df

            # Final display (use pretty weight column)
            final_display = display_df[["Wrestler", "Team", "Grade", "Level", "Weight_display", "Gender", "Matches", "Status"]]
            final_display = final_display.rename(columns={"Weight_display": "Weight"})
            # Convert only Grade and Matches to string so Streamlit left-aligns them
            final_display["Grade"] = final_display["Grade"].astype(str)
            final_display["Matches"] = final_display["Matches"].astype(str)

           # Pandas Styler to left-justify all columns while keeping numeric types
            styled = final_display.style.set_properties(**{"text-align": "left"})
            styled = styled.set_table_styles(
                [dict(selector="th", props=[("text-align", "left")])]
            )
            
            st.dataframe(styled, use_container_width=True, hide_index=True)
            # --- B: Add note for clarity ---
            st.caption("Note: Wrestlers marked as scratched are not included in this table.")

        st.markdown("---")

        st.markdown("#### Mats Overview")
        if not full_schedule:
            st.caption("No schedule yet. Go to **Match Builder** to create matchups.")
        else:
            mat_rows = []
            for m in range(1, CONFIG["NUM_MATS"] + 1):
                mat_entries = [e for e in full_schedule if e["mat"] == m]
                count = len(mat_entries)
                early_count = sum(
                    1 for e in mat_entries
                    if next(b for b in st.session_state.bout_list if b["bout_num"] == e["bout_num"])["is_early"]
                )
                mat_rows.append({
                    "Mat": m,
                    "# Bouts": count,
                    "Early Matches": early_count
                })
            st.dataframe(pd.DataFrame(mat_rows), use_container_width=True)

        st.markdown("---")

        st.markdown("#### Rest Gap Warnings")
        if not conflicts_all:
            st.caption(f"No rest conflicts detected (min gap {rest_gap} matches).")
        else:
            # Make sure slots in the warnings match the current schedule
            slot_lookup = {
                (e["mat"], e["bout_num"]): e["slot"]
                for e in full_schedule
            }
            for c in conflicts_all:
                key1 = (c["mat"], c["bout1"])
                key2 = (c["mat"], c["bout2"])
                if key1 in slot_lookup:
                    c["slot1"] = slot_lookup[key1]
                if key2 in slot_lookup:
                    c["slot2"] = slot_lookup[key2]
        
            # Build a table that only shows slots (no bout numbers)
            conflicts_df = pd.DataFrame(conflicts_all)
            conflicts_df = conflicts_df[
                ["wrestler", "team", "mat", "slot1", "slot2", "gap"]
            ].rename(columns={
                "wrestler": "Wrestler",
                "team": "Team",
                "mat": "Mat",
                "slot1": "Slot A",
                "slot2": "Slot B",
                "gap": "Gap",
            })
        
            st.warning(
                f"There are **{len(conflicts_df)}** potential rest issues "
                f"(gap < {rest_gap} matches on the same mat)."
            )
            st.dataframe(conflicts_df, use_container_width=True)

        # ==========================================================
    # TAB 3 – HELP
    # ==========================================================
    with tab_help:
        st.subheader("How to Use This Tool")

        st.markdown("##### 1. Build Your Roster CSV")
        st.markdown(
            """
Your roster CSV **must** include the following columns (in any order), and you may use  
the **Advanced options → Merge multiple roster CSV files** tool if you need to combine  
rosters from multiple teams.

### Required Columns

| Column          | Description                                          | Example      |
|-----------------|------------------------------------------------------|--------------|
| `name`          | Wrestler name                                        | `John Doe`   |
| `team`          | Team name (used for colors & legends)                | `Stillwater` |
| `grade`         | Numeric grade (5–8, etc.)                            | `7`          |
| `level`         | Level / experience (float: 1.0, 1.5, 2.0, etc.)      | `1.5`        |
| `weight`        | Weight in pounds (numeric)                           | `75`         |
| `early_match`   | `Y`/`N` – whether the wrestler needs an early match  | `Y`          |
| `scratch`       | `Y`/`N` – marked out of the meet (can be changed later) | `N`       |

### Optional Columns

| Column            | Description                                                               | Example |
|-------------------|---------------------------------------------------------------------------|---------|
| `gender`          | `M` / `F` (or similar; the app normalizes internally)                    | `F`     |
| `cross_gender_ok` | `Y`/`N` – whether this wrestler may wrestle someone of another gender    | `N`     |

You **do not** need to provide an `id` column – the app generates unique IDs automatically.

### Tips
- Use **Step 1** to download the official CSV template.
- If you receive separate roster files from multiple teams, use  
  **Advanced options → Merge multiple roster CSV files** to quickly create a single combined CSV.
- Once your roster is ready, upload it in **Step 2** to generate matchups.
"""
        )

        st.markdown("##### 2. Tune Meet Settings (Sidebar)")
        st.markdown(
            """
- **Min / Max Matches** – target range for bouts per wrestler.
- **Number of Mats** – how many mats are running at once.
- **Max Level Diff / Weight Diff** – how strict the matching is.
- **Min Rest Gap** – how many bouts must be between two matches for the same wrestler.
- **Team Colors** – after you upload a roster, each team appears here so you can assign colors used in:
  - Circle emojis in the drag rows
  - PDF/Excel exports
  - Legends in the mat previews
"""
        )

        st.markdown("##### 3. Build & Adjust the Meet")
        st.markdown(
            """
- Use **Pre-Meet Scratches** to quickly remove wrestlers from the meet.
- Use **Manual Match Creator** to fill gaps for wrestlers under the minimum and when coaches want specific bouts.
  - Wrestler 2 choices are filtered so that gender preferences are respected (cross-gender only if both wrestlers allow it).
- In **Mat Previews**:
  - Drag to reorder bouts on each mat.
  - Use the per-mat dropdown to remove a bout.
  - Use the **Move to mat** control to move a bout from one mat to another.
  - Watch the *rest warnings* and *multi-mat warnings* to avoid conflicts.
- Use the single **Undo** button to step backwards through:
  - Bout removals  
  - Drag/reorder changes  
  - Manual matches  
  - Added suggested matches  
  - Scratches updates (in edited mode)
"""
        )

        st.markdown("##### 4. Exports")
        st.markdown(
            """
- Click **Generate Coach Packets PDF** to build:
  - A **PDF** with all matches for a specific team will be generated.
- Then use **Download Coach Packet PDF** to download document.

- Click **Generate Documents** to build:
  - An **Excel** file with roster, all matchups, remaining suggestions, and mat sheets.
  - A **PDF** with mat-by-mat bout sheets, including early-match highlighting and team colors.
- Then use the **Download Excel / Download PDF** buttons to download document.
"""
        )
else:
    st.info("Upload a roster CSV in **Step 2** to unlock Match Builder, Meet Summary, and Help tabs.")

# ----------------------------------------------------------------------
# AUTOSAVE AT END OF RUN
# ----------------------------------------------------------------------
if st.session_state.get("initialized"):
    autosave_meet()
    ts = st.session_state.get("last_autosave_time")
    if ts:
        st.caption(f"💾 Autosaved this meet at {ts}.")

st.markdown("---")
st.caption("**Privacy**: Your roster is processed in your browser. Nothing is uploaded or stored.")









